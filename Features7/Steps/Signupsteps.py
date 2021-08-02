import time

from behave import *
from selenium import webdriver
from time import sleep
import os
import openpyxl
from openpyxl import workbook


@given('launch sprint1 signup chrome browser')
def launchsignupsprint1(context):
    path = str(os.getcwd())
    context.driver = webdriver.Chrome(executable_path=path + "\chromedriver.exe")
    context.driver.maximize_window()


@when('open sprint1 signup DataAxel Login')
def opensignupsprint1(context):
    context.driver.get("https://test-app.salesgenie.com/audience#n/0/")
    context.driver.find_element_by_xpath("//div[@class='filter-database-popup__overlay filter-database-popup__overlay--show']").click()
    context.driver.implicitly_wait(100)
    context.driver.find_element_by_id("thin-search-register").click()

@then('verify sprint1 that the user signup on page')
def verifysignupsprint1(context):
    Name = context.driver.find_element_by_xpath("//input[@name='name']")
    Company = context.driver.find_element_by_xpath("//input[@name='companyName']")
    Email = context.driver.find_element_by_xpath("//input[@name='email']")
    Phone = context.driver.find_element_by_xpath("//input[@name='phone']")
    password = context.driver.find_element_by_xpath("//input[@type='password']")
    Signup = context.driver.find_element_by_xpath("//button[normalize-space()='Get Started']")
    workbook = openpyxl.load_workbook("Logintest.xlsx")
    names = workbook.get_sheet_names()
    sheet = workbook.get_sheet_by_name(names[1])

    # get total Number rows
    rowCount = sheet.max_row
    colCount = sheet.max_column
    print(rowCount)
    print(colCount)

    for row in range(2, rowCount + 1):
        full_name = sheet.cell(row=row, column=1).value
        comp_name = sheet.cell(row=row, column=2).value
        email = sheet.cell(row=row, column=3).value
        reg_phone = sheet.cell(row=row, column=4).value
        reg_pass = sheet.cell(row=row, column=5).value
        Name.clear()
        Name.send_keys(full_name)
        Company.clear()
        Company.send_keys(comp_name)
        Email.clear()
        Email.send_keys(email)
        Phone.clear()
        Phone.send_keys(reg_phone)
        password.clear()
        password.send_keys(reg_pass)
        time.sleep(5)
        Signup.click()
        time.sleep(10)

    context.driver.find_element_by_xpath("//ul[@class='margin-top__1']/li[2]").click()
    context.driver.find_element_by_xpath("//button[normalize-space()='Get Code']").click()
    context.driver.find_element_by_xpath("//input[@name='verificationCode']").send_keys("123456")
    context.driver.find_element_by_xpath("//button[normalize-space()='Verify']").click()
    context.driver.get_screenshot_as_file("screen.png")

@then('sprint1signup close browser')
def closeBrowsersprint1(context):
    time.sleep(10)
    context.driver.close()
