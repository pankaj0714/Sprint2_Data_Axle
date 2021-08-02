import time

from behave import *
from selenium import webdriver
from time import sleep
import os
import openpyxl
from openpyxl import workbook


@given('launch sprint1 forgot chrome browser')
def launchBrowsersprint1(context):
    path = str(os.getcwd())
    context.driver = webdriver.Chrome(executable_path=path + "\chromedriver.exe")
    context.driver.maximize_window()


@when('open sprint1 DataAxel forgot')
def openforgotsprint1(context):
    context.driver.get("https://test-app.salesgenie.com/audience#n/0/")
    context.driver.find_element_by_xpath("//div[@class='filter-database-popup__overlay filter-database-popup__overlay--show']").click()
    context.driver.implicitly_wait(100)
    context.driver.find_element_by_id("thin-search-signin").click()
    forgetEmail = context.driver.find_element_by_xpath("//a[normalize-space()='Forgot your Password?']")
    context.driver.execute_script("arguments[0].click();", forgetEmail)
    #context.driver.find_element_by_xpath("//a[normalize-space()='Forgot your Password?']").click()


@then('verify sprint1 that the user forgot on page')
def verifyforgotsprint1(context):
    Email = context.driver.find_element_by_xpath("//input[@name='email']")
    Forget = context.driver.find_element_by_xpath("//button[normalize-space()='Submit']")
    workbook = openpyxl.load_workbook("Logintest.xlsx")
    names = workbook.get_sheet_names()
    sheet = workbook.get_sheet_by_name(names[2])

    # get total Number rows
    rowCount = sheet.max_row
    colCount = sheet.max_column
    print(rowCount)
    print(colCount)

    for row in range(2, rowCount + 1):
        email = sheet.cell(row=row, column=1).value
        Email.clear()
        Email.send_keys(email)
        time.sleep(2)
        Forget.click()
        time.sleep(5)

    context.driver.find_element_by_xpath("//ul[@class='margin-top__1']/li[3]").click()
    context.driver.find_element_by_xpath("//button[normalize-space()='Get Code']").click()
    context.driver.find_element_by_xpath("//input[@name='verificationCode']").send_keys("123456")
    context.driver.find_element_by_xpath("//button[normalize-space()='Verify']").click()
    context.driver.find_element_by_xpath("//input[@type='password']").send_keys("Kipl@123")
    context.driver.find_element_by_xpath("//button[normalize-space()='Create Password']").click()
    context.driver.implicitly_wait(100)
    context.driver.get_screenshot_as_file("screen.png")

@then('sprint1forgot close browser')
def closeBrowsersprint1(context):
    time.sleep(10)
    context.driver.close()
