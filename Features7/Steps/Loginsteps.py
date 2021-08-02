import time
from behave import *
from selenium import webdriver
import os
import openpyxl
from openpyxl import workbook


@given('launch sprint1 chrome browser')
def launchBrowsersprint1(context):
    path = str(os.getcwd())
    context.driver = webdriver.Chrome(executable_path=path + "\chromedriver.exe")
    context.driver.maximize_window()


@when('open sprint1 DataAxel signin')
def openLoginsprint1(context):
    context.driver.get("https://test-app.salesgenie.com/audience#n/0/")
    context.driver.implicitly_wait(100)
    context.driver.find_element_by_xpath("//div[@class='filter-database-popup__overlay filter-database-popup__overlay--show']").click()
    context.driver.implicitly_wait(100)
    context.driver.find_element_by_id("thin-search-signin").click()
    #context.driver.find_element_by_xpath("//button[normalize-space()='Sign In']").click()

@then('verify sprint1 that the user signin on page')
def verifyLoginsprint1(context):
    username = context.driver.find_element_by_xpath("//input[@type='text']")
    passw = context.driver.find_element_by_xpath("//input[@type='password']")
    Login = context.driver.find_element_by_xpath("//button[normalize-space()='Sign In']")
    path = str(os.getcwd())
    workbook = openpyxl.load_workbook(path+"\Logintest.xlsx")
    names = workbook.get_sheet_names()
    sheet = workbook.get_sheet_by_name(names[0])

    # get total Number rows
    rowCount = sheet.max_row
    colCount = sheet.max_column
    print(rowCount)
    print(colCount)

    for row in range(2, rowCount + 1):
        user_name = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value
        print(str(user_name) + " " + str(password))
        username.clear()
        username.send_keys(user_name)
        passw.clear()
        passw.send_keys(password)
        time.sleep(2)
        Login.click()
        time.sleep(5)

@then('sprint1login close browser')
def closeBrowsersprint1(context):
    time.sleep(10)
    context.driver.close()
